import openpyxl ##若未安装此库 cmd使用 pip install openpyxl 安装
wb = openpyxl.load_workbook('test.xlsx') ##打开目标文档
wb.create_sheet(index=1,title='操作表') ##创建操作表头
sheet = wb.worksheets[0] ## 打开第一个表头
for row in sheet.iter_rows():
	for cell in row:##遍历数据，格式自行修改
		sheet['A1'].value='cell.value'##输出，自行修改
wb.save('test.xlsx')##保存