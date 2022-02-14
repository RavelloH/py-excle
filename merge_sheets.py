import openpyxl as xls
import os
from copy import copy
from openpyxl.styles import Font

res_path = "res"
#把文件合并到源文件里，并对比不同之处标红
def merge(keywords):
    src_file = "表1-5-2教职工其他信息.xlsx"#源文件
    file_list = []
    col_num = 12
    people_dict = {}
    change_cnt =0
    print("开始合并文件\n----------------\n查找关键字:",
          keywords,"\n查找目录:",res_path)
    for root, dirs, files in os.walk(res_path):
        for file in files:
            # 获取文件所属目录
            # 获取文件路径
            file = os.path.join(root, file)
            if file.find(keywords)>0:
                file_list.append(file)
    print("发现",len(file_list),"个文件\n--------------\n进行合并")
    #获取到原文档
    wb = xls.load_workbook(src_file)
    ws = wb.worksheets[0]
    for i in range(1,ws.max_row+1):
        #先把原文件中的编号做一个映射
        uuid = ws.cell(i,1).value
        if not uuid :
            break
        people_dict[uuid] = i
    for file in file_list:
        wb_tmp = xls.load_workbook(file)
        ws_tmp = wb_tmp.worksheets[0]
        for i in range(1,ws_tmp.max_row+1):
            uuid = ws_tmp.cell(i,1).value
            row_id = 0
            is_change = False
            if not uuid:
                break
            s = "未更改"
            try:
                row_id = people_dict[uuid]
            except:
                print("未发现",uuid)
                continue
            for j in range(1,col_num+1):
                val = ws_tmp.cell(i,j).value
                if val and not val == ws.cell(row_id,j).value:
                    ws.cell(row_id,j).value = val
                    ws.cell(row_id,j).font = Font(color="33FF00")
                    change_cnt = change_cnt+1

                    is_change = True
            if is_change:
                ws.cell(row_id, col_num + 1).value = "已更改"
    wb.save(res_path+"/"+src_file)
    print("合并完成---------------")
    print(change_cnt,"处发生改变")

merge("1-5-2")
